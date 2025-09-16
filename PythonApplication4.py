# -*- coding: utf-8 -*-
"""
Serienblätter je Kreditor aus mock.xlsx
nutzt Vorlage 'Beilage Verfuegung.xlsx' und erzeugt 'Beilage_Verfuegung_per_Kreditor.xlsx'
"""

import re
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break

# === Basispfade ===
BASE_DIR = Path(r"C:\Users\peno\Beilagebrief\Beilage-Massenbrief")  ### HIER MUSS EIGENER PFAD GEWÄHLT WERDEN ###
INPUT_XLSX    = BASE_DIR / "mock.xlsx"
TEMPLATE_XLSX = BASE_DIR / "Beilage Verfuegung.xlsx"
OUTPUT_XLSX   = BASE_DIR / "Beilage_Verfuegung_per_Kreditor.xlsx"

# === Spalten in mock.xlsx ===
COL_SUP_CODE = "ithSupplierCode"
COL_SUP_NAME = "ithSupplierName"
COL_SUP_CITY = "ithSupplierCity"          # optional
COL_SUP_EXT  = "ithSupplierExternalNbr1"
COL_ER       = "ER"
COL_AMOUNT   = "itlTotalAmount"
COL_CC       = "itlCostCentreCode1"
COL_CODE     = "Code"
COL_REASON   = "Begründung"

# === Welches Register (Sheet) soll gelesen werden === 
SHEET_NAME = "Kontierung"

# === NA15: aus separates Register lesen und indizieren ===
NA15_SHEET_NAME = "NA15 Begründungen"

def _norm(s: str) -> str:
    return str(s or "").strip().casefold()

def load_na15_index(xlsx_path: Path, sheet_name: str | None = NA15_SHEET_NAME):
    """
    Liest das NA15-Register und baut ein Dict:
      key: (kreditor_name_norm, er_norm)  -> list[str]  (Begründungen)
    Erkennt gängige Spalten-Synonyme.
    """
    # Alle Sheets laden (robust gegen Reihenfolge/Hidden)
    all_sheets = pd.read_excel(xlsx_path, engine="openpyxl", sheet_name=None)

    # Explizit benanntes Sheet nehmen, sonst heuristisch suchen
    if sheet_name and sheet_name in all_sheets:
        df = all_sheets[sheet_name].copy()
    else:
        # Heuristik: erstes Blatt, das alle Pflichtspalten (oder Synonyme) enthält
        for name, df_try in all_sheets.items():
            cols_norm = [_norm(c) for c in df_try.columns]
            if any(c in cols_norm for c in ("kreditor name", "kreditor", "ithsuppliername")) \
               and any(c in cols_norm for c in ("er nr.", "er", "ernr", "er-nr")) \
               and any(c in cols_norm for c in ("begründung", "begruendung", "grund", "kommentar", "bemerkung")):
                df = df_try.copy()
                break
        else:
            # Nichts Passendes gefunden
            return {}

    # Spalten-Synonyme auf Zielnamen mappen
    colmap = {}
    for c in df.columns:
        cn = _norm(c)
        if cn in ("kreditor name", "kreditor", "ithsuppliername"):
            colmap[c] = "NA15_NAME"
        elif cn in ("er nr.", "er", "ernr", "er-nr"):
            colmap[c] = "NA15_ER"
        elif cn in ("begründung", "begruendung", "grund", "kommentar", "bemerkung"):
            colmap[c] = "NA15_REASON"
    df = df.rename(columns=colmap)

    req = {"NA15_NAME", "NA15_ER", "NA15_REASON"}
    if not req.issubset(df.columns):
        # Falls Spalten anders heißen, hier anpassen/erweitern
        return {}

    # leere Gründe raus
    df = df[df["NA15_REASON"].astype(str).str.strip() != ""].copy()

    # Index bauen: (name_norm, er_norm) -> [reasons...]
    index = {}
    for _, row in df.iterrows():
        key = (_norm(row["NA15_NAME"]), _norm(row["NA15_ER"]))
        index.setdefault(key, []).append(str(row["NA15_REASON"]).strip())
    return index

# === Vorlage-Zellen ===
CELL_SUP_CODE = "B4"   # Kreditor Nr. -> Code
CELL_SUP_NAME = "B5"   # Kreditor -> Name (ggf. Stadt)

# KORREKTE Zeilen basierend auf Vorlage-Analyse
TABLE_START_ROW = 10   # Daten starten in Zeile 10
HEADER_ROW = 8         # Header-Titel sind in Zeile 8
TEMPLATE_ROW = 9       # Diese Zeile muss gelöscht werden
TEMPLATE_NA14_ROWS = [23, 24, 25]  # Standard NA14-Bereich der Vorlage löschen

# Spaltenzuordnung basierend auf der tatsächlichen Vorlage
COLS_TEMPLATE_ORDER = [
    (COL_SUP_EXT, "A"),  # Forderungseingabe
    (COL_ER,      "B"),  # RE-Nr.
    (COL_AMOUNT,  "C"),  # Betrag in CHF
    (COL_CC,      "D"),  # Klasse (Kostenstelle gemappt)
    (COL_CODE,    "E"),  # Verfügung (Code)
    (COL_REASON,  "F"),  # Begründung
    # G = Text (bleibt leer)
]

# Kostenstellen-Legende -> Bezeichnung
COST_CENTER_MAP = {
    "9099100": "Anerkannt",
    "9099200": "Bedingt anerkannt",
    "9099300": "Bestritten",
    "9099400": "Massa",
}

def safe_sheet_name(name: str) -> str:
    if not name or str(name).strip() == "":
        name = "Sheet"
    try:
        name = str(name).encode('utf-8', errors='ignore').decode('utf-8', errors='ignore')
    except:
        name = "Sheet"
    name = re.sub(r'[\\/*?\[\]]+', "_", name)
    name = name.strip()
    return name[:31] or "Sheet"

def setup_page_formatting(ws):
    """Setzt die Seitenformatierung für A4 Querformat mit Fusszeile"""
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    ws.page_margins = PageMargins(
        left=0.7, right=0.7, top=0.75, bottom=1.0,
        header=0.3, footer=0.5
    )
    
    ws.oddFooter.center.text = "Seite &P von &N"
    ws.oddFooter.center.size = 10

def set_column_widths(ws):
    """Setzt optimale Spaltenbreiten für A4 Querformat"""
    column_widths = {
        'A': 18, 'B': 12, 'C': 15, 'D': 18, 'E': 12, 'F': 25, 'G': 15
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

def clean_template_rows(ws):
    """Löscht alle störenden Zeilen aus der Vorlage"""
    # Zeile 9: Technische Spaltennamen
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f"{col}{TEMPLATE_ROW}"].value = None
        # Auch eventuelle Rahmen/Formatierung entfernen
        ws[f"{col}{TEMPLATE_ROW}"].border = Border()
        ws[f"{col}{TEMPLATE_ROW}"].fill = PatternFill()
    
    # Zeilen 23-25: Standard NA14-Bereich aus Vorlage löschen
    for row in TEMPLATE_NA14_ROWS:
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f"{col}{row}"].value = None
            ws[f"{col}{row}"].border = Border()
            ws[f"{col}{row}"].fill = PatternFill()

def set_and_format_headers(ws, header_row):
    """Setzt die Header-Titel und formatiert sie - STRIKT nur A-F"""
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = Font(bold=True)
    bottom_border = Border(bottom=Side(style='thin'))
    
    # Nur Spalten A-F formatieren (G hat keinen Titel mehr, H nie)
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        cell = ws[f"{col}{header_row}"]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = bottom_border
    
    # Sicherstellen, dass G und H KEINE Formatierung haben
    for col in ['G', 'H']:
        cell = ws[f"{col}{header_row}"]
        cell.border = Border()  # Explizit leeren Rahmen setzen
        cell.fill = PatternFill()  # Explizit leere Füllung setzen

def apply_cell_formatting(ws, row, col_letter, value, is_total_row=False):
    """Wendet einheitliche Formatierung auf Zellen an"""
    cell = ws[f"{col_letter}{row}"]
    cell.value = value
    
    if col_letter == "A":
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    elif col_letter in ["B", "C", "D"]:
        cell.alignment = Alignment(horizontal='right', vertical='center', indent=1)
        if col_letter == "C" and isinstance(value, (int, float)) and value != 0:
            cell.number_format = "#,##0"
    elif col_letter == "E":
        cell.alignment = Alignment(horizontal='center', vertical='center')
    elif col_letter == "F":
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=1)
    else:
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    if is_total_row:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

def map_cost_center(value: str) -> str:
    """Gibt Bezeichnung aus Legende zurück; wenn unbekannt, Originalcode."""
    code = (value or "").strip()
    code_digits = "".join(ch for ch in code if ch.isdigit())
    return COST_CENTER_MAP.get(code_digits or code, COST_CENTER_MAP.get(code, code))

def calculate_optimal_na14_position(total_row_idx):
    """Berechnet die optimale Position für NA14 mit garantiertem 3-Zeilen-Abstand"""
    # IMMER 3 Zeilen Abstand nach Total-Zeile
    na14_start_row = total_row_idx + 4
    
    # Prüfen ob Seitenumbruch sinnvoll ist (ab Zeile 30)
    if na14_start_row > 30:
        print(f"NA14 wird in Zeile {na14_start_row} platziert (möglicherweise auf Seite 2)")
    
    return na14_start_row

def sort_by_c_number(df, col_name):
    """Sortiert DataFrame nach C-Nummern"""
    def extract_c_number(value):
        try:
            str_val = str(value).strip()
            if '_' in str_val:
                return int(str_val.split('_')[-1])
            elif str_val.startswith('C') and len(str_val) > 1:
                numbers = re.findall(r'\d+', str_val)
                if numbers:
                    return int(numbers[-1])
            return float('inf')
        except:
            return float('inf')
    
    df_sorted = df.copy()
    df_sorted['_sort_key'] = df_sorted[col_name].apply(extract_c_number)
    df_sorted = df_sorted.sort_values('_sort_key').drop('_sort_key', axis=1)
    return df_sorted

def main():
    if not INPUT_XLSX.exists():
        raise FileNotFoundError(f"Eingabedatei fehlt: {INPUT_XLSX}")
    if not TEMPLATE_XLSX.exists():
        raise FileNotFoundError(f"Vorlage fehlt: {TEMPLATE_XLSX}")

    try:
        df = pd.read_excel(INPUT_XLSX, engine="openpyxl", sheet_name=SHEET_NAME)
        na15_index = load_na15_index(INPUT_XLSX, NA15_SHEET_NAME)

    except UnicodeDecodeError:
        df = pd.read_excel(INPUT_XLSX, engine="openpyxl", encoding="latin1")

    required = [COL_SUP_CODE, COL_SUP_NAME, COL_SUP_EXT, COL_ER, COL_AMOUNT, COL_CC, COL_CODE, COL_REASON]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Pflichtspalten fehlen in {INPUT_XLSX.name}: {missing}")

    # Normalisieren
    for c in [COL_SUP_CODE, COL_SUP_NAME, COL_SUP_CITY, COL_SUP_EXT, COL_ER, COL_CC, COL_CODE, COL_REASON]:
        if c in df.columns:
            df[c] = df[c].astype(str).fillna("").apply(
                lambda x: x.encode('utf-8', errors='ignore').decode('utf-8', errors='ignore').strip()
            )
    df[COL_AMOUNT] = pd.to_numeric(df[COL_AMOUNT], errors="coerce").fillna(0.0)

    wb = load_workbook(TEMPLATE_XLSX)
    base_ws = wb.active
    base_title = base_ws.title

    sup_cols = [COL_SUP_CODE, COL_SUP_NAME] + ([COL_SUP_CITY] if COL_SUP_CITY in df.columns else [])
    suppliers = (
        df[sup_cols]
        .drop_duplicates(subset=[COL_SUP_CODE])
        .sort_values(by=[COL_SUP_NAME, COL_SUP_CODE])
        .to_dict(orient="records")
    )

    for sup in suppliers:
        code = sup.get(COL_SUP_CODE, "")
        name = sup.get(COL_SUP_NAME, "")
        city = sup.get(COL_SUP_CITY, "") if COL_SUP_CITY in sup else ""

        part = df[df[COL_SUP_CODE] == code].copy()
        part = sort_by_c_number(part, COL_SUP_EXT)

        ws = wb.copy_worksheet(base_ws)
        ws.title = safe_sheet_name(name or code or "Kreditor")

        setup_page_formatting(ws)
        set_column_widths(ws)

        ws[CELL_SUP_CODE] = code
        ws[CELL_SUP_NAME] = f"{name}{(', ' + city) if city else ''}"

        # WICHTIG: Alle störenden Vorlage-Zeilen löschen
        clean_template_rows(ws)
        
        # Header formatieren (strikt nur A-G)
        set_and_format_headers(ws, HEADER_ROW)

        # Datenzeilen
        start_row = TABLE_START_ROW
        total_amount_sheet = float(part[COL_AMOUNT].sum())
        
        for i, (_, row) in enumerate(part.iterrows(), start=0):
            r = start_row + i
            for col_name, col_letter in COLS_TEMPLATE_ORDER:
                val = row.get(col_name, "")
                if col_name == COL_CC:
                    val = map_cost_center(val)
                elif col_name == COL_AMOUNT:
                    val = float(row.get(col_name, 0))
                apply_cell_formatting(ws, r, col_letter, val, is_total_row=False)

        # Total-Zeile (ohne Spalte G zu formatieren)
        total_row_idx = start_row + len(part)
        for col_letter, val in [("A", "Total"), ("B", ""), ("C", total_amount_sheet), 
                               ("D", ""), ("E", ""), ("F", "")]:
            apply_cell_formatting(ws, total_row_idx, col_letter, val, is_total_row=True)
        
        # Spalte G in Total-Zeile explizit NICHT formatieren
        ws[f"G{total_row_idx}"].fill = PatternFill()  # Keine Füllung
        ws[f"G{total_row_idx}"].border = Border()     # Kein Rahmen

        
        # --- NA15-Begründungen (aus separatem Register) unterhalb einfügen ---
        # ERs dieses Kreditors, die NA15 in der Haupttabelle haben:
        try:
            ers_na15 = (
                part.loc[part[COL_CODE].str.upper() == "NA15", COL_ER]
                    .astype(str).str.strip()
                    .dropna()
                    .unique().tolist()
            )
        except Exception:
            ers_na15 = []

        # Passende Begründungen aus dem separaten NA15-Register holen
        rows = []
        name_norm = _norm(name)  # Kreditor-Name aus dem aktuellen Blatt
        for er in sorted(ers_na15, key=lambda x: x):
            key = (name_norm, _norm(er))
            reasons = na15_index.get(key, [])
            if reasons:
                # mehrere Gründe pro ER ggf. zusammenfassen
                rows.append((er, "\n\n".join(reasons)))

        if rows:
            # Startposition: 3 Leerzeilen unter der Totals-Zeile
            block_start = calculate_optimal_na14_position(total_row_idx)  # liefert total_row_idx + 4

            # Überschrift
            ws[f"A{block_start}"] = "Begründungen (NA15)"
            ws[f"A{block_start}"].font = Font(bold=True, size=12)
            ws[f"A{block_start}"].alignment = Alignment(horizontal='left', vertical='top')

            # Kopfzeile
            hdr = block_start + 1
            ws[f"A{hdr}"] = "ER Nr."
            ws[f"B{hdr}"] = "Begründung"
            header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            header_font = Font(bold=True)
            bottom_border = Border(bottom=Side(style='thin'))
            for col in ["A", "B"]:
                c = ws[f"{col}{hdr}"]
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = bottom_border

            # Optional: Begründungs-Spalte breiter machen (B..F zusammenführen)
            try:
                ws.merge_cells(start_row=hdr, start_column=2, end_row=hdr, end_column=6)  # B..F
            except Exception:
                pass

            r = hdr + 1
            for er_val, reason_text in rows:
                ws[f"A{r}"] = er_val
                ws[f"A{r}"].alignment = Alignment(horizontal='center', vertical='top')

                ws[f"B{r}"] = reason_text
                ws[f"B{r}"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

                # Merge pro Datenzeile (B..F) – falls gewünscht
                try:
                    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
                except Exception:
                    pass

                # Zeilenhöhe grob nach Textlänge
                if reason_text:
                    est_lines = max(1, len(reason_text) // 80 + reason_text.count("\n") + 1)
                    ws.row_dimensions[r].height = min(est_lines * 15, 180)

                r += 1

            print(f"    → NA15-Begründungen hinzugefügt: {len(rows)} Einträge")
        else:
            # Keine NA15-Fälle oder keine Begründungen im NA15-Sheet für diesen Kreditor/ER
            pass


    wb.remove(wb[base_title])
    wb.save(OUTPUT_XLSX)
    print(f"Fertig. Datei erstellt:\n{OUTPUT_XLSX}")

if __name__ == "__main__":

    main()
