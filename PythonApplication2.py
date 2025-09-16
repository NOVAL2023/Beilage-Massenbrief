# -*- coding: utf-8 -*-
"""
Serienblätter je Kreditor aus mock.xlsx - Verbesserte Version mit Fehlerbehandlung
"""

import re
import sys
from pathlib import Path
import traceback

print("=== SCHULDENRUF VERFÜGUNG GENERATOR ===")
print(f"Python Version: {sys.version}")
print()

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.worksheet.pagebreak import Break
    print("✓ Alle Module erfolgreich importiert")
except ImportError as e:
    print(f"✗ FEHLER beim Importieren der Module: {e}")
    print("Installieren Sie fehlende Module mit: pip install pandas openpyxl")
    sys.exit(1)

# === Basispfade ===
BASE_DIR = Path(r"C:\Users\peno\Beilagebrief\Beilage-Massenbrief")
INPUT_XLSX = BASE_DIR / "mock.xlsx"
TEMPLATE_XLSX = BASE_DIR / "Beilage Verfuegung.xlsx"
OUTPUT_XLSX = BASE_DIR / "Beilage_Verfuegung_per_Kreditor.xlsx"

print(f"Arbeitsverzeichnis: {BASE_DIR}")
print(f"Eingabedatei: {INPUT_XLSX}")
print(f"Vorlage: {TEMPLATE_XLSX}")
print(f"Ausgabedatei: {OUTPUT_XLSX}")
print()

# === Dateien prüfen ===
if not BASE_DIR.exists():
    print(f"✗ FEHLER: Basisverzeichnis nicht gefunden: {BASE_DIR}")
    sys.exit(1)
else:
    print(f"✓ Basisverzeichnis gefunden")

if not INPUT_XLSX.exists():
    print(f"✗ FEHLER: Eingabedatei nicht gefunden: {INPUT_XLSX}")
    sys.exit(1)
else:
    print(f"✓ Eingabedatei gefunden ({INPUT_XLSX.stat().st_size} Bytes)")

if not TEMPLATE_XLSX.exists():
    print(f"✗ FEHLER: Vorlage nicht gefunden: {TEMPLATE_XLSX}")
    sys.exit(1)
else:
    print(f"✓ Vorlage gefunden ({TEMPLATE_XLSX.stat().st_size} Bytes)")

print()

# Rest des ursprünglichen Codes (Konstanten)
COL_SUP_CODE = "ithSupplierCode"
COL_SUP_NAME = "ithSupplierName"
COL_SUP_CITY = "ithSupplierCity"
COL_SUP_EXT = "ithSupplierExternalNbr1"
COL_ER = "ER"
COL_AMOUNT = "itlTotalAmount"
COL_CC = "itlCostCentreCode1"
COL_CODE = "Code"
COL_REASON = "Begründung"

CELL_SUP_CODE = "B4"
CELL_SUP_NAME = "B5"

TABLE_START_ROW = 10
HEADER_ROW = 8
TEMPLATE_ROW = 9
TEMPLATE_NA14_ROWS = [23, 24, 25]

COLS_TEMPLATE_ORDER = [
    (COL_SUP_EXT, "A"),
    (COL_ER, "B"),
    (COL_AMOUNT, "C"),
    (COL_CC, "D"),
    (COL_CODE, "E"),
    (COL_REASON, "F"),
]

COST_CENTER_MAP = {
    "9099100": "Anerkannt",
    "9099200": "Bedingt anerkannt",
    "9099300": "Bestritten",
    "9099400": "Massa",
}

def safe_sheet_name(name: str) -> str:
    if not name or str(name).strip() == "":
        name = "Sheet"
    try:
        name = str(name).encode('utf-8', errors='ignore').decode('utf-8', errors='ignore')
    except:
        name = "Sheet"
    name = re.sub(r'[\\/*?\[\]]+', "_", name)
    name = name.strip()
    return name[:31] or "Sheet"

def setup_page_formatting(ws):
    try:
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        
        ws.page_margins = PageMargins(
            left=0.7, right=0.7, top=0.75, bottom=1.0,
            header=0.3, footer=0.5
        )
        
        ws.oddFooter.center.text = "Seite &P von &N"
        ws.oddFooter.center.size = 10
    except Exception as e:
        print(f"Warnung bei Seitenformatierung: {e}")

def set_column_widths(ws):
    try:
        column_widths = {
            'A': 18, 'B': 12, 'C': 15, 'D': 18, 'E': 12, 'F': 25, 'G': 15
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
    except Exception as e:
        print(f"Warnung bei Spaltenbreiten: {e}")

def clean_template_rows(ws):
    try:
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f"{col}{TEMPLATE_ROW}"].value = None
            ws[f"{col}{TEMPLATE_ROW}"].border = Border()
            ws[f"{col}{TEMPLATE_ROW}"].fill = PatternFill()
        
        for row in TEMPLATE_NA14_ROWS:
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                ws[f"{col}{row}"].value = None
                ws[f"{col}{row}"].border = Border()
                ws[f"{col}{row}"].fill = PatternFill()
    except Exception as e:
        print(f"Warnung beim Bereinigen der Vorlage: {e}")

def set_and_format_headers(ws, header_row):
    try:
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        header_font = Font(bold=True)
        bottom_border = Border(bottom=Side(style='thin'))
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            cell = ws[f"{col}{header_row}"]
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = bottom_border
        
        for col in ['G', 'H']:
            cell = ws[f"{col}{header_row}"]
            cell.border = Border()
            cell.fill = PatternFill()
    except Exception as e:
        print(f"Warnung bei Header-Formatierung: {e}")

def apply_cell_formatting(ws, row, col_letter, value, is_total_row=False):
    try:
        cell = ws[f"{col_letter}{row}"]
        cell.value = value
        
        if col_letter == "A":
            cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        elif col_letter in ["B", "C", "D"]:
            cell.alignment = Alignment(horizontal='right', vertical='center', indent=1)
            if col_letter == "C" and isinstance(value, (int, float)) and value != 0:
                cell.number_format = "#,##0"
        elif col_letter == "E":
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif col_letter == "F":
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=1)
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        
        if is_total_row:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    except Exception as e:
        print(f"Warnung bei Zellformatierung: {e}")

def map_cost_center(value: str) -> str:
    code = (value or "").strip()
    code_digits = "".join(ch for ch in code if ch.isdigit())
    return COST_CENTER_MAP.get(code_digits or code, COST_CENTER_MAP.get(code, code))

def calculate_optimal_na14_position(total_row_idx):
    na14_start_row = total_row_idx + 4
    if na14_start_row > 30:
        print(f"  → NA14 wird in Zeile {na14_start_row} platziert (möglicherweise auf Seite 2)")
    return na14_start_row

def sort_by_c_number(df, col_name):
    def extract_c_number(value):
        try:
            str_val = str(value).strip()
            if '_' in str_val:
                return int(str_val.split('_')[-1])
            elif str_val.startswith('C') and len(str_val) > 1:
                numbers = re.findall(r'\d+', str_val)
                if numbers:
                    return int(numbers[-1])
            return float('inf')
        except:
            return float('inf')
    
    df_sorted = df.copy()
    df_sorted['_sort_key'] = df_sorted[col_name].apply(extract_c_number)
    df_sorted = df_sorted.sort_values('_sort_key').drop('_sort_key', axis=1)
    return df_sorted

def main():
    try:
        print("=== DATENVERARBEITUNG STARTET ===")
        
        # Excel-Datei einlesen
        print("Lade Excel-Datei...")
        try:
            df = pd.read_excel(INPUT_XLSX, engine="openpyxl")
            print(f"✓ {len(df)} Zeilen eingelesen")
        except UnicodeDecodeError:
            print("Versuche mit latin1 encoding...")
            df = pd.read_excel(INPUT_XLSX, engine="openpyxl", encoding="latin1")
            print(f"✓ {len(df)} Zeilen eingelesen (latin1)")
        except Exception as e:
            print(f"✗ FEHLER beim Einlesen der Excel-Datei: {e}")
            return False

        # Spalten prüfen
        print("Prüfe Spalten...")
        required = [COL_SUP_CODE, COL_SUP_NAME, COL_SUP_EXT, COL_ER, COL_AMOUNT, COL_CC, COL_CODE, COL_REASON]
        missing = [c for c in required if c not in df.columns]
        if missing:
            print(f"✗ FEHLER: Pflichtspalten fehlen: {missing}")
            print(f"Verfügbare Spalten: {list(df.columns)}")
            return False
        print("✓ Alle Pflichtspalten gefunden")

        # Daten normalisieren
        print("Normalisiere Daten...")
        for c in [COL_SUP_CODE, COL_SUP_NAME, COL_SUP_CITY, COL_SUP_EXT, COL_ER, COL_CC, COL_CODE, COL_REASON]:
            if c in df.columns:
                df[c] = df[c].astype(str).fillna("").apply(
                    lambda x: x.encode('utf-8', errors='ignore').decode('utf-8', errors='ignore').strip()
                )
        df[COL_AMOUNT] = pd.to_numeric(df[COL_AMOUNT], errors="coerce").fillna(0.0)
        print(f"✓ Daten normalisiert")

        # Vorlage laden
        print("Lade Vorlage...")
        try:
            wb = load_workbook(TEMPLATE_XLSX)
            base_ws = wb.active
            base_title = base_ws.title
            print(f"✓ Vorlage geladen: {base_title}")
        except Exception as e:
            print(f"✗ FEHLER beim Laden der Vorlage: {e}")
            return False

        # Kreditoren verarbeiten
        sup_cols = [COL_SUP_CODE, COL_SUP_NAME] + ([COL_SUP_CITY] if COL_SUP_CITY in df.columns else [])
        suppliers = (
            df[sup_cols]
            .drop_duplicates(subset=[COL_SUP_CODE])
            .sort_values(by=[COL_SUP_NAME, COL_SUP_CODE])
            .to_dict(orient="records")
        )
        
        print(f"Verarbeite {len(suppliers)} Kreditoren...")
        
        for i, sup in enumerate(suppliers, 1):
            code = sup.get(COL_SUP_CODE, "")
            name = sup.get(COL_SUP_NAME, "")
            city = sup.get(COL_SUP_CITY, "") if COL_SUP_CITY in sup else ""
            
            print(f"  {i}/{len(suppliers)}: {name} ({code})")
            
            try:
                part = df[df[COL_SUP_CODE] == code].copy()
                part = sort_by_c_number(part, COL_SUP_EXT)

                ws = wb.copy_worksheet(base_ws)
                ws.title = safe_sheet_name(name or code or "Kreditor")

                setup_page_formatting(ws)
                set_column_widths(ws)

                ws[CELL_SUP_CODE] = code
                ws[CELL_SUP_NAME] = f"{name}{(', ' + city) if city else ''}"

                clean_template_rows(ws)
                set_and_format_headers(ws, HEADER_ROW)

                start_row = TABLE_START_ROW
                total_amount_sheet = float(part[COL_AMOUNT].sum())
                
                for j, (_, row) in enumerate(part.iterrows(), start=0):
                    r = start_row + j
                    for col_name, col_letter in COLS_TEMPLATE_ORDER:
                        val = row.get(col_name, "")
                        if col_name == COL_CC:
                            val = map_cost_center(val)
                        elif col_name == COL_AMOUNT:
                            val = float(row.get(col_name, 0))
                        apply_cell_formatting(ws, r, col_letter, val, is_total_row=False)

                total_row_idx = start_row + len(part)
                for col_letter, val in [("A", "Total"), ("B", ""), ("C", total_amount_sheet), 
                                       ("D", ""), ("E", ""), ("F", "")]:
                    apply_cell_formatting(ws, total_row_idx, col_letter, val, is_total_row=True)
                
                ws[f"G{total_row_idx}"].fill = PatternFill()
                ws[f"G{total_row_idx}"].border = Border()

                na14 = part[part[COL_CODE].str.upper() == "NA14"]
                na14_texts = [t for t in na14[COL_REASON].astype(str) if t.strip()]
                if na14_texts:
                    na14_row = calculate_optimal_na14_position(total_row_idx)
                    
                    ws[f"A{na14_row}"] = "Begründung (Anderes/Rechtsstreit)"
                    ws[f"A{na14_row}"].font = Font(bold=True, size=12)
                    ws[f"A{na14_row}"].alignment = Alignment(horizontal='left', vertical='top')
                    
                    combined_text = "\n\n".join(na14_texts)
                    ws[f"A{na14_row + 1}"] = combined_text
                    ws[f"A{na14_row + 1}"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    
                    estimated_lines = max(2, len(combined_text) // 80)
                    ws.row_dimensions[na14_row + 1].height = min(estimated_lines * 15, 150)
                    
                    print(f"    → NA14 Begründung hinzugefügt")

            except Exception as e:
                print(f"    ✗ FEHLER bei Kreditor {name}: {e}")
                continue

        # Speichern
        print("Speichere Datei...")
        try:
            wb.remove(wb[base_title])
            wb.save(OUTPUT_XLSX)
            print(f"✓ Datei erfolgreich erstellt: {OUTPUT_XLSX}")
            return True
        except Exception as e:
            print(f"✗ FEHLER beim Speichern: {e}")
            return False

    except Exception as e:
        print(f"✗ UNERWARTETER FEHLER: {e}")
        traceback.print_exc()
        return False

if __name__ == "__main__":
    try:
        success = main()
        if success:
            print("\n=== ERFOLGREICH ABGESCHLOSSEN ===")
        else:
            print("\n=== FEHLER AUFGETRETEN ===")
            sys.exit(1)
    except KeyboardInterrupt:
        print("\n=== ABGEBROCHEN ===")
        sys.exit(1)
    except Exception as e:
        print(f"\n=== KRITISCHER FEHLER: {e} ===")
        traceback.print_exc()
        sys.exit(1)